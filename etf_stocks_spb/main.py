from openpyxl import load_workbook, Workbook
import json
import os
import tinvest as tin
import requests
from fake_useragent import UserAgent
from bs4 import BeautifulSoup
import concurrent.futures


HEADER = {'User-Agent': str(UserAgent().chrome)}


class Stock:
    def __init__(self, ticker, isin='blank'):
        self.ticker = ticker
        self.isin = isin
        self.short = False
        self.risk = [100, 100]
        self.atr = 1
        self.price = 1
        self.beta = 1
        self.atr_to_price = 1
        self.sector = ''
        self.industry = ''
        self.avg_volume = 1


class Sheet:
    def __init__(self, name='default'):
        self.name = name
        self.etf_title = ['Ticker', 'Weight', 'SPB', 'Short', 'ATR', 'Beta', 'Price', 'Avg volume']
        self.spb_title = ['Ticker', 'Name', 'Sector', 'Industry', 'Country']
        self.rows = []


def get_market_stocks(client):
    payload = client.get_market_stocks().payload
    stocks_usd = [stock for stock in payload.instruments[:] if stock.currency == 'USD']
    stocks = {s.ticker: Stock(ticker=s.ticker, isin=s.isin) for s in stocks_usd}
    return stocks


def get_data(stocks_dict, f_path, sheets_list):
    wb = load_workbook(filename=f_path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet = Sheet(sheetname)

        for row in ws.rows:
            if row[0].value == "Ticker":
                continue
            spb = False
            if row[0].value in stocks_dict:
                spb = True
            sheet.rows.append([row[0].value, row[1].value, spb])

        sheets_list.append(sheet)


class Liquidity:
    def __init__(self, short_able, risk_long, risk_short):
        self.short_able = short_able
        self.risk_long = risk_long
        self.risk_short = risk_short


def check_liquidity(stocks):
    url = "https://www.tinkoff.ru/invest/margin/equities/"
    r = requests.get(url, headers=HEADER)
    soup = BeautifulSoup(r.content, 'lxml')

    liquidity = {}

    body_table = soup.find('table').find('tbody')
    trs = body_table.find_all('tr')
    for tr in trs:
        tds = tr.find_all('td')
        isin = tds[1].text
        if tds[2].text == "Доступен":
            short_able = True
        else:
            short_able = False
        risk_long = float(tds[3].text.split('/')[0])
        risk_short = float(tds[3].text.split('/')[1])

        liquidity[isin] = Liquidity(short_able, risk_long, risk_short)

    for s in stocks:
        try:
            l = liquidity[s.isin]
            s.short = l.short_able
            s.risk = [l.risk_long, l.risk_short]
        except KeyError:
            continue


def add_short_info(sheets_list, stocks_dict):
    url = "https://www.tinkoff.ru/invest/margin/equities/"
    r = requests.get(url, headers=HEADER)
    soup = BeautifulSoup(r.content, 'lxml')

    isin_able = {}

    all_tr = soup.find_all('tr', class_='Table__row_3Unlc Table__row_clickable_3EeUg')
    for tr in all_tr:
        all_td = tr.find_all('td')
        isin = all_td[1].text
        ability = all_td[2].text
        if ability == "Доступен":
            isin_able[isin] = True
        else:
            isin_able[isin] = False

    for sheet in sheets_list:
        for row in sheet.rows:
            if row[2]:
                try:
                    row.append(isin_able[stocks_dict[row[0]].isin])
                except KeyError:
                    row.append(False)


def check_hash(s):
    if s == '-':
        return 0
    return float(s)


def get_soup(url):
    r = requests.get(url, headers=HEADER)
    return BeautifulSoup(r.content, 'lxml')


def get_finviz_table(soup: BeautifulSoup):
    table_mult = soup.find('table', class_='snapshot-table2')
    snapshot_td2_cp = table_mult.find_all('td', class_='snapshot-td2-cp')
    snapshot_td2 = table_mult.find_all('td', class_='snapshot-td2')
    mult = {snapshot_td2_cp[i].text.lower(): snapshot_td2[i].text.lower()
            for i in range(len(snapshot_td2))}
    return mult


def get_finviz_info_spb(ticker):
    soup = get_soup(f"https://finviz.com/quote.ashx?t={ticker}")
    title = soup.find('table', class_="fullview-title")
    try:
        all_links = title.find_all(class_='tab-link')
        name = all_links[0].text
        sector = all_links[1].text
        industry = all_links[2].text
        country = all_links[3].text
        return [ticker, name, sector, industry, country]
    except:
        return [ticker, 0, 0, 0, 0]


def get_avg_volume(avg):
    if avg[-1] == 'm':
        avg = float(avg[0:-1]) * 1_000_000
    elif avg[-1] == 'k':
        avg = float(avg[0:-1]) * 1_000
    else:
        avg = float(avg[0:-1])
    return avg


def stock_finviz_info(stock: Stock):
    soup = get_soup(f"https://finviz.com/quote.ashx?t={stock.ticker}")
    mult = get_finviz_table(soup)
    stock.atr = check_hash(mult['atr'])
    stock.price = check_hash(mult['price'])
    stock.beta = check_hash(mult['beta'])
    stock.atr_to_price = stock.atr / stock.price
    stock.avg_volume = get_avg_volume(mult['avg volume'])

    links = soup.find('table', class_='fullview-title').find_all(class_='tab-link')
    stock.sector = links[1].text
    stock.industry = links[2].text


def get_finviz_info(rows):
    for row in rows:
        if not row[2]:
            continue

        soup = get_soup(f"https://finviz.com/quote.ashx?t={row[0]}")
        mult = get_finviz_table(soup)

        row.append(check_hash(mult['atr']))
        row.append(check_hash(mult['beta']))
        row.append(check_hash(mult['price']))
        row.append(get_avg_volume(mult['avg volume']))


def add_finviz_info(sheets_list):
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for sheet in sheets_list:
            executor.submit(get_finviz_info, sheet.rows)
    # get_finviz_info(sheets_list[0].rows)


def write_sheet_etf(sheets_list, f_path):
    wb = Workbook()
    ws = wb.active
    ws.title = sheets_list[0].name

    for i in range(len(sheets_list)):
        if i != 0:
            ws = wb.create_sheet(sheets_list[i].name)

        title = sheets_list[i].etf_title
        rows = sheets_list[i].rows

        for cal in range(len(title)):
            ws.cell(column=cal+1, row=1, value=title[cal])

        for row in range(len(rows)):
            for cal in range(len(rows[row])):
                ws.cell(column=cal+1, row=row+2, value=rows[row][cal])

    wb.save(f_path)


def get_stocks():
    with open(os.path.join(os.path.expanduser('~'), 'no_commit', 'info.json')) as f:
        data = json.load(f)
    client = tin.SyncClient(data['token_tinkoff_real'])
    return get_market_stocks(client)


def etf_stocks():
    all_stocks = get_stocks()
    my_sheets = []

    get_data(all_stocks, "etfs.xlsx", my_sheets)
    
    add_short_info(my_sheets, all_stocks)
    add_finviz_info(my_sheets)

    write_sheet_etf(my_sheets, "etfs_holdings.xlsx")


def margin_stocks(f_path):
    stocks = list(get_stocks().values())
    check_liquidity(stocks)
    stocks = [s for s in stocks if s.risk[0] < 25]
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(stock_finviz_info, s) for s in stocks}

    title = ['Ticker', 'Short', 'Price', 'ATR', 'ATR to Price', 'Beta', 'Long risk', 'Short risk',
             'Avg Volume', 'Sector', 'Industry']
    rows = [[s.ticker, s.short, s.price, s.atr, s.atr_to_price, s.beta,
             s.risk[0], s.risk[1], s.avg_volume, s.sector, s.industry]
            for s in stocks if s.price > 1]
    col_num = len(title)
    wb = Workbook()
    ws = wb.active

    for cal in range(col_num):
        ws.cell(column=cal+1, row=1, value=title[cal])

    for row in range(len(rows)):
        for cal in range(col_num):
            ws.cell(column=cal+1, row=row+2, value=rows[row][cal])

    wb.save(f_path)

def spb_stocks(f_path):
    tickers = list(get_stocks().keys())
    sheet = Sheet()
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(get_finviz_info_spb, ticker) for ticker in tickers}
        for f in concurrent.futures.as_completed(futures):
            sheet.rows.append(f.result())

    wb = Workbook()
    ws = wb.active
    col_number = len(sheet.spb_title)

    for cal in range(col_number):
        ws.cell(column=cal+1, row=1, value=sheet.spb_title[cal])

    for row in range(len(sheet.rows)):
        for cal in range(col_number):
            ws.cell(column=cal+1, row=row+2, value=sheet.rows[row][cal])

    wb.save(f_path)


def main():
    # etf_stocks()
    # spb_stocks('E:\Document\spb_stocks.xlsx')
    margin_stocks('E:\Document\margin_stocks.xlsx')


if __name__ == '__main__':
    main()
