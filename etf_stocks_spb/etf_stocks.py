from market import get_stocks
from openpyxl import load_workbook, Workbook
from models import Sheet
from scrap import add_short_info, get_soup, get_finviz_table, check_hash, get_avg_volume
import concurrent.futures


def get_finviz_info(rows):
    for row in rows:
        if not row[2]:
            continue

        soup = get_soup(f"https://finviz.com/quote.ashx?t={row[0]}")
        mult = get_finviz_table(soup)

        row.append(check_hash(mult['atr']))
        row.append(check_hash(mult['beta']))
        row.append(check_hash(mult['price']))
        row.append(get_avg_volume(mult['avg volume']))


def add_finviz_info(sheets_list):
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        for sheet in sheets_list:
            executor.submit(get_finviz_info, sheet.rows)


def write_sheet_etf(sheets_list, f_path):
    wb = Workbook()
    ws = wb.active
    ws.title = sheets_list[0].name

    for i in range(len(sheets_list)):
        if i != 0:
            ws = wb.create_sheet(sheets_list[i].name)

        title = sheets_list[i].etf_title
        rows = sheets_list[i].rows

        for cal in range(len(title)):
            ws.cell(column=cal+1, row=1, value=title[cal])

        for row in range(len(rows)):
            for cal in range(len(rows[row])):
                ws.cell(column=cal+1, row=row+2, value=rows[row][cal])

    wb.save(f_path)


def get_data(stocks_dict, f_path, sheets_list):
    wb = load_workbook(filename=f_path)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        sheet = Sheet(sheetname)

        for row in ws.rows:
            if row[0].value == "Ticker":
                continue
            spb = False
            if row[0].value in stocks_dict:
                spb = True
            sheet.rows.append([row[0].value, row[1].value, spb])

        sheets_list.append(sheet)


def etf_stocks():
    all_stocks = get_stocks()
    my_sheets = []

    get_data(all_stocks, "etfs.xlsx", my_sheets)

    add_short_info(my_sheets, all_stocks)
    add_finviz_info(my_sheets)

    write_sheet_etf(my_sheets, "etfs_holdings.xlsx")


etf_stocks()