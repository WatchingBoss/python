from openpyxl import Workbook
import concurrent.futures
from models import Stock
from market import get_stocks
from scrap import get_soup, get_finviz_table, check_hash, check_liquidity, get_avg_volume


def stock_finviz_info(stock: Stock):
    soup = get_soup(f"https://finviz.com/quote.ashx?t={stock.ticker}")
    mult = get_finviz_table(soup)
    stock.atr = check_hash(mult['atr'])
    stock.price = check_hash(mult['price'])
    stock.beta = check_hash(mult['beta'])
    stock.atr_to_price = stock.atr / stock.price
    stock.avg_volume = get_avg_volume(mult['avg volume'])

    links = soup.find('table', class_='fullview-title').find_all(class_='tab-link')
    stock.sector = links[1].text
    stock.industry = links[2].text


def margin_stocks(f_path):
    stocks = list(get_stocks().values())
    check_liquidity(stocks)
    stocks = [s for s in stocks if s.risk[0] < 25]
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(stock_finviz_info, s) for s in stocks}

    title = ['Ticker', 'Short', 'Price', 'ATR', 'ATR to Price', 'Beta', 'Long risk', 'Short risk',
             'Avg Volume', 'Sector', 'Industry']
    rows = [[s.ticker, s.short, s.price, s.atr, s.atr_to_price, s.beta,
             s.risk[0], s.risk[1], s.avg_volume, s.sector, s.industry]
            for s in stocks if s.price > 1]
    col_num = len(title)
    wb = Workbook()
    ws = wb.active

    for cal in range(col_num):
        ws.cell(column=cal+1, row=1, value=title[cal])

    for row in range(len(rows)):
        for cal in range(col_num):
            ws.cell(column=cal+1, row=row+2, value=rows[row][cal])

    wb.save(f_path)


margin_stocks('E:\Document\margin_stocks.xlsx')
