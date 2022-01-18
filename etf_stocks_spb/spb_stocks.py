from market import get_stocks
from openpyxl import Workbook
from models import Sheet
from scrap import get_soup
import concurrent.futures


def get_finviz_info_spb(ticker):
    soup = get_soup(f"https://finviz.com/quote.ashx?t={ticker}")
    title = soup.find('table', class_="fullview-title")
    try:
        all_links = title.find_all(class_='tab-link')
        name = all_links[0].text
        sector = all_links[1].text
        industry = all_links[2].text
        country = all_links[3].text
        return [ticker, name, sector, industry, country]
    except:
        return [ticker, 0, 0, 0, 0]


def spb_stocks(f_path):
    tickers = list(get_stocks().keys())
    sheet = Sheet()
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as executor:
        futures = {executor.submit(get_finviz_info_spb, ticker) for ticker in tickers}
        for f in concurrent.futures.as_completed(futures):
            sheet.rows.append(f.result())

    wb = Workbook()
    ws = wb.active
    col_number = len(sheet.spb_title)

    for cal in range(col_number):
        ws.cell(column=cal+1, row=1, value=sheet.spb_title[cal])

    for row in range(len(sheet.rows)):
        for cal in range(col_number):
            ws.cell(column=cal+1, row=row+2, value=sheet.rows[row][cal])

    wb.save(f_path)


spb_stocks('E:\Document\spb_stocks.xlsx')